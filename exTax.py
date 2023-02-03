'''
Created on 2021/02/06
見せばやな雄島のあまの袖だにも
　濡れにぞ濡れし色は変わらず
　　　　　　殷富門院大輔
@author: sue
'''

from typing import NamedTuple
import openpyxl as xl
import csv
import os
import jaconv

import c
import d
import e


'''
勘定科目内訳明細書用のExcelファイルを読込み、
e-Tax送信用のCSVファイルを作成する。

設定.csvファイルの内容に従い処理する。

仕様
必要科目は設定ファイルを参照する
Excelファイルに指定されたシートがなければ、無視
タイトル行の削除
シートを結合する
空白行は先頭のセルが空で判断
　Excelの操作によっては空白行もデータありとしてしまうことがある
必要な項目を超えた右側のセルは無視
合計行が使えない場合は、合計行を削除
金額にカンマがあっても対応（何もしていないが、そうなっている）

文字数、不要箇所などのエラーチェックはしない
設定.csvで、
　半角→全角に修正する列番号、
　３０文字で切捨てる列番号　を指定
'''

class UchiwakeSheet(NamedTuple):
    kubun : object  # tuple of int    (14, 1)
    kamoku : str    # '役員給与等（代表者）'
    sheet_mei : str # '区分「14-1」HOI141_5.0_役員給与等（代表者）'
    csv_file_mei : str  # 'HOI141_5.0_役員給与等'
    title_gyou_suu : int    # 削除するタイトル行数
    koumoku_suu : int   # 項目数
    goukei_umu : bool   # 合計行を認めるか否か
    gyou_kubun_ichi : int   # 行区分の項目の位置　通常は2
    shiyou : bool   # 使用しているか否か
    kana_retsu : object   # tuple of int 半角文字が含まれる可能性のある列
    kirisute_30 : object    # tuple of int ３０文字で切捨てる列


def read_settei_file(file_name):
    '''
    設定ＣＳＶファイルを読込む。

    Parameters
    ----------
    file_mame : str
        読込む設定CSVファイル名。

    Returns
    -------
    version : str
        設定ファイルのバージョン名
        ex. 平成31年4月1日以後終了事業年度分
    bunrui : str
        分類名
        ex. 標準, 〇〇会社
    excel_file_mame : str
        勘定科目内訳明細書の内容が記載されているExcelファイル名。
    uchiwake_sheet_list : list of UchiwakeSheet
        Excelファイルの各シートの内容
    '''
    d.dprint_method_start()
    with open(file_name) as f:
        reader = csv.reader(f)
        l = [row for row in reader]
    version = l[0][1]
    bunrui = l[1][1]
    excel_file_name = l[2][1]
    csv_folder_name = l[3][1]
    uchiwake_sheet_list = []
    for index in range(5, len(l)):
        if (l[index][1] == None) or (l[index][1] == ''):
            kubun = (int(l[index][0]),)
        else:
            kubun = (int(l[index][0]), int(l[index][1]))
        goukei_umu = True if l[index][7] == '1' else False
        shiyou = True if l[index][9] == '1' else False
        uchiwake = UchiwakeSheet(
                kubun,
                l[index][2],
                l[index][3],
                l[index][4],
                int(l[index][5]),
                int(l[index][6]),
                goukei_umu,
                int(l[index][8]),
                shiyou,
                (int(l[index][10]), int(l[index][11]), int(l[index][12])),   # 20210823
                (int(l[index][13]),)
                )
        uchiwake_sheet_list.append(uchiwake)
        d.dprint(uchiwake)
    d.dprint_method_end()
    return (version, bunrui, excel_file_name, csv_folder_name,
            uchiwake_sheet_list)


def read_excel(excel_file_name, csv_folder_name, uchiwake_sheet_list):
    '''
    ExcelファイルからＣＳＶファイルを作成する。

    Parameters
    ----------
    excel_file_mame : str
        読込むExcelファイル名。
    csv_folder_mame : str
        作成するCSVファイルを保存するフォルダ名
    uchiwake_sheet_list : list of UchiwakeSheet
        Excelファイルの各シートの内容

    Returns
    -------
    kekka : bool
        成功か、失敗か。
    '''
    d.dprint_method_start()
    tandoku_list = []
    fukusuu_list = []
    sudeni_list = []
    for index, uchiwake in enumerate(uchiwake_sheet_list):
        if index in sudeni_list:
            continue
        if not uchiwake.shiyou:
            continue
        index_list = [index]
        for index_fukusuu, uchiwake_fukusuu \
                in enumerate(uchiwake_sheet_list[index+1:]):
#             if uchiwake.csv_file_mei == uchiwake_fukusuu.csv_file_mei:
            if uchiwake.kubun[0] == uchiwake_fukusuu.kubun[0]:
                if uchiwake_fukusuu.shiyou:
                    index_list.append(index_fukusuu+index+1)
                    sudeni_list.append(index_fukusuu+index+1)
        if len(index_list) == 1:
            tandoku_list.append(index)
        else:
            fukusuu_list.append(index_list)

    book = xl.load_workbook(excel_file_name, data_only=True)
    d.dprint("単独シート")
    for tandoku in tandoku_list:
        d.dprint(tandoku)
        uchiwake = uchiwake_sheet_list[tandoku]
        if uchiwake.sheet_mei \
                not in book.sheetnames:
            d.dprint("{}シートがないので飛ばす".format(uchiwake.sheet_mei))
            continue
        d.dprint(uchiwake)
        sheet = book[uchiwake.sheet_mei]
        csv_file_name = os.path.join(csv_folder_name,
                uchiwake.csv_file_mei+'.csv')
        sheet_to_csv(sheet, csv_file_name,
            uchiwake.title_gyou_suu, uchiwake.koumoku_suu,
            uchiwake.goukei_umu, uchiwake.gyou_kubun_ichi,
            uchiwake.kana_retsu, uchiwake.kirisute_30)    # 20210823
    d.dprint("複数シート")
    for fukusuu in fukusuu_list:
        d.dprint(fukusuu)
        sheet_list = []
        settei_list = []
        for index in fukusuu:
            uchiwake = uchiwake_sheet_list[index]
            if uchiwake.sheet_mei \
                    not in book.sheetnames:
                continue
            sheet_list.append(book[uchiwake.sheet_mei])
            settei_list.append((uchiwake.title_gyou_suu,
                    uchiwake.koumoku_suu,
                    uchiwake.goukei_umu,
                    uchiwake.gyou_kubun_ichi,
                    uchiwake.kana_retsu,    # 20210823
                    uchiwake.kirisute_30))   # 20210823
        csv_file_name = os.path.join(csv_folder_name,
                    uchiwake.csv_file_mei+'.csv')
        sheet_fukusuu_to_csv(sheet_list, csv_file_name, settei_list)
    d.dprint_method_end()


def sheet_to_csv(sheet, csv_file_name,
            title_gyou_suu, koumoku_suu, goukei_umu, gyou_kubun_ichi,
            kana_retsu, kirisute_30):
    '''
    Excelの単独のシートからＣＳＶファイルを作成する。

    Parameters
    ----------
    sheet : worksheet
        Excelのシート。
    csv_file_mame : str
        作成するCSVファイル名。
    title_gyou_suu : int
        削除するタイトルの行数。
    koumoku_suu : int
        項目数。
    goukei_umu : bool
        合計行を認めるか否か
    gyou_kubun_ichi : int
        行区分の項目の位置　通常は2
    kana_retsu : tuple of int
        半角文字が含まれている可能性のある列　左端(A)が1、0なら無し
    kirisute_30 : tuple of int
        半角文字が含まれている可能性のある列　左端(A)が1、0なら無し

    Returns
    -------
    kekka : bool
        成功か、失敗か。
    '''
    d.dprint_method_start()
    csv_data = []
    sheet_to_data(sheet, csv_data,
            title_gyou_suu, koumoku_suu, goukei_umu, gyou_kubun_ichi,
            kana_retsu, kirisute_30)
    write_csv_file(csv_file_name, csv_data)
    d.dprint_method_end()

def sheet_fukusuu_to_csv(sheet_list, csv_file_name,
            settei_list):
    '''
    Excelの複数のシートからＣＳＶファイルを作成する。

    Parameters
    ----------
    sheet_list : list of worksheet
        Excelのシートのリスト。
    csv_file_mame : str
        作成するCSVファイル名。
    settei_list : list of tuple of int, int, bool, tuple of int, tuple of int
        title_gyou_suu : int
            削除するタイトルの行数。
        koumoku_suu : int
            項目数。
        goukei_umu : bool
            合計行を認めるか否か
        gyou_kubun_ichi : int
            行区分の項目の位置　通常は2
        kana_retsu : tuple of int
            半角カナが含まれている可能性のある列　左端(A)が1、0なら無し
        kirisute_30 : tuple of int
            半角文字が含まれている可能性のある列　左端(A)が1、0なら無し

    Returns
    -------
    kekka : bool
        成功か、失敗か。
    '''
    d.dprint_method_start()
    csv_data = []
    for (sheet, settei) in zip(sheet_list, settei_list):
        sheet_to_data(sheet, csv_data,
            settei[0], settei[1], settei[2], settei[3],
            settei[4], settei[5])
    write_csv_file(csv_file_name, csv_data)
    d.dprint_method_end()

def sheet_to_data(sheet, data_list,
            title_gyou_suu, koumoku_suu, goukei_umu, gyou_kubun_ichi,
            kana_retsu, kirisute_30):
    '''
    Excelの単独のシートからデータのリストを作成する。

    Parameters
    ----------
    sheet : worksheet
        Excelのシート。
    data_list : list of list of str
        CSVファイルに書き込むデータのリスト。
    title_gyou_suu : int
        削除するタイトルの行数。
    koumoku_suu : int
        項目数。
    goukei_umu : bool
        合計行を認めるか否か
    gyou_kubun_ichi : int
        行区分の項目の位置　通常は2
    kana_retsu : tuple of int
        半角カナが含まれている可能性のある列　左端(A)が1、0なら無し
    kirisute_30 : tuple of int
        半角文字が含まれている可能性のある列　左端(A)が1、0なら無し

    Returns
    -------
    kekka : bool
        成功か、失敗か。
    '''
    d.dprint_method_start()
    d.dprint(sheet)
#     d.dprint_name("kana_retsu", kana_retsu)
    for row in range(title_gyou_suu + 1, sheet.max_row + 1):
        if not goukei_umu:
            if sheet.cell(row=row, column=gyou_kubun_ichi).value \
                    == '1':
                # 合計行を認めない内訳書なので、合計行を飛ばす
                d.dprint("合計行を飛ばす")
                continue
        if sheet.cell(row=row, column=1).value == None:
            d.dprint("空の行なので終了")
            break
        row_data = []
        for column in range(1, koumoku_suu + 1):
            # 20210823
#             d.dprint(column)
#             d.dprint(sheet.cell(row=row, column=column).value)
            data = sheet.cell(row=row, column=column).value
#             d.dprint(data)
            if (column in kana_retsu) and (data is not None):
#                 d.dprint(sheet.cell(row=row, column=column).value)
                data = jaconv.h2z(data)
                data = jaconv.h2z(data, kana=False, ascii=True, digit=False)
            if type(data) is str:
                data = data.replace('\n', ' ')
            # 20210823
#             d.dprint(kirisute_30)
            if (column in kirisute_30) and (data is not None):
#                 d.dprint(data)
                data = data[:30]
#                 d.dprint(data)
            # 20220201
            if data is None:
                data = ""
            row_data.append(data)
        data_list.append(row_data)
    d.dprint_method_end()

def write_csv_file(csv_file_name, data_list):
    '''
    データリストからＣＳＶファイルを作成する。

    Parameters
    ----------
    csv_file_mame : str
        作成するCSVファイル名。
    data_list : list of list of object
        CSVファイルに出力するデータ。

    Returns
    -------
    '''
    d.dprint_method_start()
    d.dprint(csv_file_name)
    d.dprint(data_list)
    with open(csv_file_name, "w", newline='') as f:
        writer = csv.writer(f)
        for row in data_list:
            d.dprint(row)
            writer.writerow(row)
    d.dprint_method_end()


if __name__ == '__main__':
    (version, bunrui, excel_file_name, csv_folder_name,
            uchiwake_sheet_list) = read_settei_file(
                    "設定.csv")
    read_excel(excel_file_name, csv_folder_name, uchiwake_sheet_list)


